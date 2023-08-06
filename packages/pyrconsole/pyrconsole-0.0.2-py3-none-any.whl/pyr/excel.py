import ustring         as s
from   ulog            import log

import cons            as cons
import ulog            as ulog
import fileio          as fileio
import tar             as tar
import stock           as stock


def translate(m):
    m00 = m
    m = kuiup(m)
    ulog.log_trans(m00, m)
    return m


def kuiup(m):
    if (is_kuiup(m)):
        m = "excel " + m
    return m


def is_kuiup(m):
    return m == "kuiup"


def handle(m):
    if (is_kuiup(m)):
        do_kuiup()


def do_kuiup():
    f = tar.rp("dd/Kui.xlsx")
    import openpyxl
    wb = openpyxl.load_workbook(f)
    sheet_n = wb.sheetnames[0]
    sheet = wb[sheet_n]
    sheet.cell(row=3, column=5).value = stock.get_stock_price("xmjt")
    sheet.cell(row=4, column=5).value = stock.get_stock_price("txkg")
    wb.save(f)

